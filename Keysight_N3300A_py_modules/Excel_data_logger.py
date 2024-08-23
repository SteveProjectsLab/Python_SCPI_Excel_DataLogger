import time
import os
import datetime

from openpyxl import Workbook, load_workbook

class Excel_data_logger:

    m_v=0.005455112#0.005412529     #slope of V     - from sperimental caracterization 
    q_v=-0.01695449#0.001975504     #intercept of V - from sperimental caracterization

    m_i=0.032264347     #(old)0.03238         #slope of I     - from sperimental caracterization 
    q_i=-65.09716032    #(old)-65.48128       #intercept of I - from sperimental caracterization
    
    #builder to create a new log file
    def __init__(self):
        self.fileName=str("log_"+datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")+".xlsx")
        self.filePath=os.path.join(os.getcwd())
        self.wb=Workbook()
        self.ws=self.wb.active
        
        self.ws['A1'].value= "s_num"    #sample number
        self.ws['B1'].value= "time"     #sample log time
        self.ws['C1'].value= "Eload_I [A]"  #I from load
        self.ws['D1'].value= "Eload_V [V]"  #V from load
        self.ws['E1'].value= "CH0"      #ADC CH0 from stm32
        self.ws['F1'].value= "CH1"      #ADC CH1 from stm32
        self.ws['G1'].value= "sys_I [A]"    #I from stm32
        self.ws['H1'].value= "sys_V [V]"    #V from stm32
        self.ws['I1'].value= "mult [V]"    #V from stm32
    def storeDataELN3300A(self,curr,tens,mycurr,mytens,i):
        i=i+2
        if(i==2):
            print(str(self.ws['A1'].value)+'\t'+
                  str(self.ws['B1'].value)+'\t\t'+
                  str(self.ws['C1'].value)+'\t'+
                  str(self.ws['D1'].value)+'\t'+
                  str(self.ws['E1'].value)+'\t'+
                  str(self.ws['F1'].value)+'\t'+
                  str(self.ws['G1'].value)+'\t'+
                  str(self.ws['H1'].value)
                  )
        print(str(str(i-1)+
                  '\t'+datetime.datetime.now().strftime("%H-%M-%S")+
                  '\t'+curr+
                  '\t'+tens+
                  '\t'+mycurr+
                  '\t'+mytens+
                  '\t'+str(round(float(mycurr)*self.m_i+self.q_i,4))+
                  '\t\t'+str(round(float(mytens)*self.m_v+self.q_v,4))))

        self.fileName=str("log_eload_stm_"+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
        
        continua_misure = True
        self.ws['A'+str(i)].value= float(i-1)
        self.ws['B'+str(i)].value= datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        self.ws['C'+str(i)].value= float(curr)    #load's current measurement
        self.ws['D'+str(i)].value= float(tens)     #load's tension measurement
        self.ws['E'+str(i)].value= float(mycurr)  #stm32 CH0 ADC value
        self.ws['F'+str(i)].value= float(mytens)  #stm32 CH1 ADC value
        self.ws['G'+str(i)].value= float(mycurr)*self.m_i+self.q_i #stm32 current measurement
        self.ws['H'+str(i)].value= float(mytens)*self.m_v+self.q_v #stm32 tension measurement
        
    def storeDataSTM32(self,mycurr,mytens,i):
        i=i+2
        if(i==2):
            print(str(self.ws['A1'].value)+'\t'+
                  str(self.ws['B1'].value)+'\t\t'+
                  str(self.ws['E1'].value)+'\t'+
                  str(self.ws['F1'].value)+'\t'+
                  str(self.ws['G1'].value)+'\t'+
                  str(self.ws['H1'].value)
                  )
        print(str(str(i-1)+
                  '\t'+datetime.datetime.now().strftime("%H-%M-%S")+
                  '\t'+mycurr+
                  '\t'+mytens+
                  '\t'+str(round(float(mycurr)*self.m_i+self.q_i,4))+
                  '\t\t'+str(round(float(mytens)*self.m_v+self.q_v,4))))

        self.fileName=str("log_stm_"+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
        
        continua_misure = True
        self.ws['A'+str(i)].value= int(i-1)
        self.ws['B'+str(i)].value= datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        self.ws['C'+str(i)].value= "disconnected"    
        self.ws['D'+str(i)].value= "disconnected"     
        self.ws['E'+str(i)].value= float(mycurr)  #stm32 CH0 ADC value
        self.ws['F'+str(i)].value= float(mytens)  #stm32 CH1 ADC value
        self.ws['G'+str(i)].value= float(mycurr)*self.m_i+self.q_i #stm32 current measurement
        self.ws['H'+str(i)].value= float(mytens)*self.m_v+self.q_v #stm32 tension measurement

    def storeDataSTM32DM(self,mycurr,mytens,mult_tens,i):
        i=i+2
        if(i==2):
            print(str(self.ws['A1'].value)+'\t'+
                  str(self.ws['B1'].value)+'\t\t'+
                  str(self.ws['E1'].value)+'\t'+
                  str(self.ws['F1'].value)+'\t'+
                  str(self.ws['G1'].value)+'\t'+
                  str(self.ws['H1'].value)+'\t'+
                  str(self.ws['I1'].value)
                  )
        print(str(str(i-1)+
                  '\t'+datetime.datetime.now().strftime("%H-%M-%S")+
                  '\t'+mycurr+
                  '\t'+mytens+
                  '\t'+str(round(float(mycurr)*self.m_i+self.q_i,4))+
                  '\t\t'+str(round(float(mytens)*self.m_v+self.q_v,4))+
                  '\t\t'+str(float(mult_tens))
                  ))

        self.fileName=str("log_stm_DM_"+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
        
        continua_misure = True
        self.ws['A'+str(i)].value= int(i-1)
        self.ws['B'+str(i)].value= datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        self.ws['C'+str(i)].value= "disconnected"    
        self.ws['D'+str(i)].value= "disconnected"     
        self.ws['E'+str(i)].value= float(mycurr)  #stm32 CH0 ADC value
        self.ws['F'+str(i)].value= float(mytens)  #stm32 CH1 ADC value
        self.ws['G'+str(i)].value= float(mycurr)*self.m_i+self.q_i #stm32 current measurement
        self.ws['H'+str(i)].value= float(mytens)*self.m_v+self.q_v #stm32 tension measurement
        self.ws['I'+str(i)].value= float(mult_tens) #multimeter tension measurement

    def storeDataELN3300A_DM_STM32(self,curr,tens,mult_tens,mycurr,mytens,i):
        i=i+2
        if(i==2):
            self.ws['I1'].value="mult_tension [V]"
            print(str(self.ws['A1'].value)+'\t'+
                  str(self.ws['B1'].value)+'\t\t'+
                  str(self.ws['C1'].value)+'\t'+
                  str(self.ws['D1'].value)+'\t'+
                  str(self.ws['E1'].value)+'\t'+
                  str(self.ws['F1'].value)+'\t'+
                  str(self.ws['G1'].value)+'\t'+
                  str(self.ws['H1'].value)+'\t'+
                  str(self.ws['I1'].value)
                  )
        print(str(str(i-1)+
                  '\t'+datetime.datetime.now().strftime("%H-%M-%S")+
                  '\t'+str(curr)+
                  '\t'+str(tens)+
                  '\t'+str(mycurr)+
                  '\t'+str(mytens)+
                  '\t'+str(round(float(mycurr)*self.m_i+self.q_i,4))+
                  '\t\t'+str(round(float(mytens)*self.m_v+self.q_v,4))+
                  '\t\t'+str(float(mult_tens))))

        self.fileName=str("log_eload_DM_stm_"+datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")+".xlsx")
        
        continua_misure = True
        self.ws['A'+str(i)].value= float(i-1)
        self.ws['B'+str(i)].value= datetime.datetime.now().strftime("%m-%d-%Y_%H-%M-%S")
        self.ws['C'+str(i)].value= float(curr)    #load's current measurement
        self.ws['D'+str(i)].value= float(tens)     #load's tension measurement
        self.ws['E'+str(i)].value= float(mycurr)  #stm32 CH0 ADC value
        self.ws['F'+str(i)].value= float(mytens)  #stm32 CH1 ADC value
        self.ws['G'+str(i)].value= float(mycurr)*self.m_i+self.q_i #stm32 current measurement
        self.ws['H'+str(i)].value= float(mytens)*self.m_v+self.q_v #stm32 tension measurement
        self.ws['I'+str(i)].value= float(mult_tens) #multimeter tension measurement

    def save(self):
        print('\nsaving the file...\n')
        self.wb.save(self.fileName)
